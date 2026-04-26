import streamlit as st
import io
import PyPDF2
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

# Regex para caracteres ilegales en Excel
ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

def clean_for_excel(text):
    """Elimina caracteres ilegales para Excel y espacios extra"""
    if not text: return ""
    text = str(text)
    text = ILLEGAL_CHARACTERS_RE.sub("", text)
    return text.strip()

def procesar_santander_rio_prueba(archivo_pdf, cuits_propios=None):
    """Procesa archivos PDF de Santander Rio con Estilo Dashboard Multi-Moneda + Hojas Ingresos/Egresos"""
    if cuits_propios is None:
        cuits_propios = []
    st.info("Procesando archivo de Santander Rio (Prueba)...")

    try:
        # Reinicializar el archivo para lectura
        archivo_pdf.seek(0)
        
        # Abrir el PDF usando PyPDF2
        reader = PyPDF2.PdfReader(io.BytesIO(archivo_pdf.read()))
        texto_completo = "".join(page.extract_text() + "\n" for page in reader.pages)
        
        lineas_raw = texto_completo.splitlines()

        # 1. Metadatos (Titular, Periodo)
        titular_global = "Sin Especificar"
        periodo_global = "Sin Especificar"
        
        # Titular: Linea anterior a "CUIT:" o "CUIL:"
        for i, l in enumerate(lineas_raw[:20]):
            if "CUIT:" in l or "CUIL:" in l:
                if i > 0:
                    titular_global = lineas_raw[i-1].strip()
                break
        
        # Periodo: "Desde: 27/01/23" ... "Hasta: 02/03/23"
        f_desde = None
        f_hasta = None
        for l in lineas_raw[:30]:
            match_d = re.search(r"Desde:\s*(\d{2}/\d{2}/\d{2,4})", l)
            if match_d: f_desde = match_d.group(1)
            match_h = re.search(r"Hasta:\s*(\d{2}/\d{2}/\d{2,4})", l)
            if match_h: f_hasta = match_h.group(1)
        
        if f_desde and f_hasta:
            periodo_global = f"Del {f_desde} al {f_hasta}"

        # --- DELIMITAR SECCIONES ---
        idx_pesos = None
        idx_dolares = None
        idx_fin_pesos = None # Fin de pesos puede ser inicio dolares o fin documento
        idx_fin_dolares = None

        for i, l in enumerate(lineas_raw):
            if "Movimientos en pesos" in l and idx_pesos is None:
                idx_pesos = i
            if "Movimientos en dólares" in l and idx_dolares is None:
                idx_dolares = i
            if ("Así usaste tu dinero este mes" in l or "Detalle impositivo" in l) and idx_fin_dolares is None and idx_dolares is not None:
                idx_fin_dolares = i
            # Si no hay dolares, el fin de pesos puede ser "Así usaste..."
            if ("Así usaste tu dinero este mes" in l or "Detalle impositivo" in l) and idx_fin_pesos is None and idx_pesos is not None and idx_dolares is None:
                idx_fin_pesos = i

        # Ajustar rangos
        lineas_pesos = []
        lineas_dolares = []

        if idx_pesos is not None:
            # Fin de pesos es idx_dolares si existe, sino idx_fin_pesos, sino fin archivo
            end_p = idx_dolares if idx_dolares is not None else (idx_fin_pesos if idx_fin_pesos is not None else len(lineas_raw))
            lineas_pesos = lineas_raw[idx_pesos+1 : end_p]
        
        if idx_dolares is not None:
            end_d = idx_fin_dolares if idx_fin_dolares is not None else len(lineas_raw)
            lineas_dolares = lineas_raw[idx_dolares+1 : end_d]

        # --- FUNCION EXTRACTION (REUTILIZADA) ---
        def extraer_datos_seccion(lineas):
            movimientos_text = []
            linea_actual = ""
            saldo_ini = 0.0
            saldo_fin = 0.0
            
            # Pre-procesado para unir líneas
            for l in lineas:
                # Filtrar encabezados repetidos de página (ej: "2 -  11")
                if re.match(r'^\s*\d+\s*-\s+\d+\s*$', l.strip()):
                    continue
                # Filtrar encabezados de tabla repetidos
                if "Cuenta Corriente" in l and ("CBU:" in l or "Nº" in l):
                    continue
                if "FechaComprobante" in l or "FechaComprobanteMovimiento" in l:
                    continue

                # Extraer saldos si aparecen en la sección
                if "Saldo Inicial" in l:
                    matches = re.findall(r"(-?)\$\s?([\d\.]+,\d{2})|(-?)U\$S\s?([\d\.]+,\d{2})", l)
                    # matches devuelve tuplas con grupos vacios, hay que filtrar
                    for m in matches:
                        # m = ('-', '1.200,00', '', '') para pesos
                        # m = ('', '', '-', '100,00') para dolares
                        val_str = m[1] if m[1] else m[3]
                        sign_str = m[0] if m[1] else m[2]
                        if val_str:
                            try:
                                num = float(val_str.replace(".", "").replace(",", "."))
                                if sign_str == "-": num *= -1
                                saldo_ini = num
                            except: pass

                if "Saldo total" in l:
                    matches = re.findall(r"(-?)\$\s?([\d\.]+,\d{2})|(-?)U\$S\s?([\d\.]+,\d{2})", l)
                    for m in matches:
                        val_str = m[1] if m[1] else m[3]
                        sign_str = m[0] if m[1] else m[2]
                        if val_str:
                            try: 
                                num = float(val_str.replace(".", "").replace(",", "."))
                                if sign_str == "-": num *= -1
                                saldo_fin = num
                            except: pass
                    continue # NO unir la linea de Saldo Total al movimiento anterior

                # Unir lineas de movimientos
                if re.match(r"\d{2}/\d{2}/\d{2}", l):
                    if linea_actual: movimientos_text.append(linea_actual.strip())
                    linea_actual = l
                else:
                    linea_actual += " " + l
            if linea_actual: movimientos_text.append(linea_actual.strip())

            # Parsear Movimientos usando diferencia de saldos
            parsed_data = []
            saldo_anterior = saldo_ini  # Arrancar con el saldo inicial
            
            for mov in movimientos_text:
                if "Movimientos en" in mov: continue 
                if "Saldo Inicial" in mov: continue # Filtrar Saldo Inicial siempre
                
                fecha = mov[:8]
                resto = mov[8:]
                
                # Limpieza basica de moneda para facilitar regex unico
                mov_clean = mov.replace("U$S", "$").replace("U$s", "$")

                # Buscamos todos los montos monetarios
                montos = re.findall(r"([+-]?\$\s*[\d\.,]+)", mov_clean)
                
                importe = 0.0
                desc = ""
                
                if len(montos) >= 2:
                    # El ÚLTIMO monto es el saldo acumulado (balance running)
                    str_saldo = montos[-1]
                    clean_saldo = str_saldo.replace("$", "").replace("+", "").replace("-", "").strip().replace(".", "").replace(",", ".")
                    signo_saldo = -1 if "-" in str_saldo else 1
                    try:
                        saldo_actual = float(clean_saldo) * signo_saldo
                    except:
                        saldo_actual = saldo_anterior
                    
                    # Importe = diferencia de saldos (positivo = crédito, negativo = débito)
                    importe = round(saldo_actual - saldo_anterior, 2)
                    saldo_anterior = saldo_actual
                    
                    # Descripción: todo lo que hay antes del penúltimo monto
                    str_imp = montos[-2]
                    idx_imp = mov_clean.rfind(str_imp) 
                    if idx_imp != -1:
                        desc = mov[:idx_imp] # Incluye fecha en los primeros 8 chars
                        desc = desc[8:].strip() # Quitar fecha
                    else:
                        desc = resto

                    # Limpieza extra: Quitar numeros pegados al inicio (ej: 77367269Transferencia)
                    desc = re.sub(r'^\d+', '', desc).strip()

                    parsed_data.append((fecha, clean_for_excel(desc), importe, mov))

                elif len(montos) == 1:
                    # Solo hay un monto, puede ser saldo inicial o algo raro.
                    if "Saldo Inicial" in mov: continue 
                    # Si es un movimiento sin saldo acumulado visible? Raro en este banco.
                    pass
            
            return parsed_data, saldo_ini, saldo_fin

        # Procesar
        datos_pesos, saldo_ini_pesos, saldo_fin_pesos = extraer_datos_seccion(lineas_pesos)
        datos_dolares, saldo_ini_dolares, saldo_fin_dolares = extraer_datos_seccion(lineas_dolares)

        # --- GENERACIÓN EXCEL MULTI-HOJA ---
        output = io.BytesIO()
        wb = Workbook()
        # Eliminar hoja default
        wb.remove(wb.active)
        
        # Estilos
        color_bg_main = "EC0000" 
        color_txt_main = "FFFFFF"
        thin_border = Border(left=Side(style='thin', color="A6A6A6"), right=Side(style='thin', color="A6A6A6"), top=Side(style='thin', color="A6A6A6"), bottom=Side(style='thin', color="A6A6A6"))
        fill_head_deb = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        fill_col_deb = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
        fill_row_deb = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
        fill_head_cred = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        fill_col_cred = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
        fill_row_cred = PatternFill(start_color="F2F9F1", end_color="F2F9F1", fill_type="solid")
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        red_font = Font(color='9C0006', bold=True)

        def crear_hoja_dashboard(wb, nombre_hoja, datos, s_ini, s_fin, formato_moneda='"$ "#,##0.00'):
            ws = wb.create_sheet(title=nombre_hoja)
            ws.sheet_view.showGridLines = False
            
            # Ignorar columna raw_text si existe (4 columnas)
            if datos and len(datos[0]) == 4:
                df = pd.DataFrame(datos, columns=["Fecha", "Descripcion", "Importe", "_raw"])
                df = df.drop(columns=["_raw"])
            else:
                df = pd.DataFrame(datos, columns=["Fecha", "Descripcion", "Importe"])
            
            creditos = df[df["Importe"] > 0].copy()
            debitos = df[df["Importe"] < 0].copy()
            debitos["Importe"] = debitos["Importe"].abs() # Positivo para mostrar
            
            if df.empty:
                creditos = pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])
                debitos = pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])

            # Header
            ws.merge_cells("A1:G1")
            tit = ws["A1"]
            tit.value = f"REPORTE SANTANDER ({nombre_hoja}) - {clean_for_excel(titular_global)}"
            tit.font = Font(size=14, bold=True, color=color_txt_main)
            tit.fill = PatternFill(start_color=color_bg_main, end_color=color_bg_main, fill_type="solid")
            tit.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 25

            # Metadata
            ws["A3"] = "SALDO INICIAL"
            ws["A3"].font = Font(bold=True, size=10, color="666666")
            ws["B3"] = s_ini
            ws["B3"].number_format = formato_moneda
            ws["B3"].font = Font(bold=True, size=11)
            ws["B3"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

            ws["A4"] = "SALDO FINAL"
            ws["A4"].font = Font(bold=True, size=10, color="666666")
            ws["B4"] = s_fin
            ws["B4"].number_format = formato_moneda
            ws["B4"].font = Font(bold=True, size=11)
            ws["B4"].border = Border(bottom=Side(style='thin', color="DDDDDD"))
            
            ws["D3"] = "TITULAR"; 
            ws.merge_cells("E3:G3"); ws["E3"] = clean_for_excel(titular_global)
            ws["E3"].alignment = Alignment(horizontal='center')

            ws["D4"] = "PERÍODO"; 
            ws.merge_cells("E4:G4"); ws["E4"] = clean_for_excel(periodo_global)
            ws["E4"].alignment = Alignment(horizontal='center')
            
            ws["D6"] = "CONTROL DE SALDOS"
            
            # Control Formula Placeholder
            ws["D7"] = 0
            ws["D7"].font = Font(bold=True, size=12); ws["D7"].border = thin_border
            ws.conditional_formatting.add('D7', CellIsRule(operator='notEqual', formula=['0'], stopIfTrue=True, fill=red_fill, font=red_font))

            # Tablas
            f_header = 10
            # Creditos
            ws.merge_cells(f"A{f_header}:C{f_header}"); ws[f"A{f_header}"] = "CRÉDITOS"
            ws[f"A{f_header}"].fill = fill_head_cred; ws[f"A{f_header}"].font = Font(bold=True, color="FFFFFF")
            ws[f"A{f_header}"].alignment = Alignment(horizontal="center", vertical="center")
            # Debitos
            ws.merge_cells(f"E{f_header}:G{f_header}"); ws[f"E{f_header}"] = "DÉBITOS"
            ws[f"E{f_header}"].fill = fill_head_deb; ws[f"E{f_header}"].font = Font(bold=True, color="FFFFFF")
            ws[f"E{f_header}"].alignment = Alignment(horizontal="center", vertical="center")
            
            # Subheaders
            for col, txt in zip(["A","B","C", "E","F","G"], ["Fecha","Descripción","Importe", "Fecha","Descripción","Importe"]):
                ws[f"{col}{f_header+1}"] = txt
                ws[f"{col}{f_header+1}"].border = thin_border
                ws[f"{col}{f_header+1}"].alignment = Alignment(horizontal='center')
                if col in ["A","B","C"]: ws[f"{col}{f_header+1}"].fill = fill_col_cred
                else: ws[f"{col}{f_header+1}"].fill = fill_col_deb
            
            # Llenar Creditos
            row = f_header + 2
            start_cred = row
            if creditos.empty:
                ws[f"A{row}"] = "SIN MOVIMIENTOS"; ws.merge_cells(f"A{row}:C{row}")
                ws[f"A{row}"].alignment = Alignment(horizontal='center'); ws[f"A{row}"].font = Font(italic=True, color="666666")
                row += 1
            else:
                for _, r in creditos.iterrows():
                    ws[f"A{row}"] = r["Fecha"]; ws[f"B{row}"] = r["Descripcion"]; ws[f"C{row}"] = r["Importe"]
                    ws[f"C{row}"].number_format = formato_moneda
                    for c in ["A","B","C"]: ws[f"{c}{row}"].border = thin_border; ws[f"{c}{row}"].fill = fill_row_cred
                    row += 1
            
            total_cred_row = row
            ws.merge_cells(f"A{total_cred_row}:B{total_cred_row}")
            ws[f"A{total_cred_row}"] = "TOTAL CRÉDITOS"
            ws[f"A{total_cred_row}"].font = Font(bold=True); ws[f"A{total_cred_row}"].alignment = Alignment(horizontal='right')
            ws[f"C{total_cred_row}"] = f"=SUM(C{start_cred}:C{total_cred_row-1})"
            ws[f"C{total_cred_row}"].number_format = formato_moneda; ws[f"C{total_cred_row}"].font = Font(bold=True)
            for c in ["A","B","C"]: ws[f"{c}{total_cred_row}"].border = thin_border
            
            # Llenar Debitos
            row = f_header + 2
            start_deb = row
            if debitos.empty:
                ws[f"E{row}"] = "SIN MOVIMIENTOS"; ws.merge_cells(f"E{row}:G{row}")
                ws[f"E{row}"].alignment = Alignment(horizontal='center'); ws[f"E{row}"].font = Font(italic=True, color="666666")
                row += 1
            else:
                for _, r in debitos.iterrows():
                    ws[f"E{row}"] = r["Fecha"]; ws[f"F{row}"] = r["Descripcion"]; ws[f"G{row}"] = r["Importe"]
                    ws[f"G{row}"].number_format = formato_moneda
                    for c in ["E","F","G"]: ws[f"{c}{row}"].border = thin_border; ws[f"{c}{row}"].fill = fill_row_deb
                    row += 1
            
            total_deb_row = row
            ws.merge_cells(f"E{total_deb_row}:F{total_deb_row}")
            ws[f"E{total_deb_row}"] = "TOTAL DÉBITOS"
            ws[f"E{total_deb_row}"].font = Font(bold=True); ws[f"E{total_deb_row}"].alignment = Alignment(horizontal='right')
            ws[f"G{total_deb_row}"] = f"=SUM(G{start_deb}:G{total_deb_row-1})"
            ws[f"G{total_deb_row}"].number_format = formato_moneda; ws[f"G{total_deb_row}"].font = Font(bold=True)
            for c in ["E","F","G"]: ws[f"{c}{total_deb_row}"].border = thin_border

            # Update Control Formula final
            ws["D7"] = f"=ROUND(B3+C{total_cred_row}-G{total_deb_row}-B4, 2)"
            ws["D7"].number_format = formato_moneda
            
            # Anchos
            ws.column_dimensions["B"].width = 40; ws.column_dimensions["F"].width = 40
            ws.column_dimensions["C"].width = 18; ws.column_dimensions["G"].width = 18

        # =====================================================
        # CATEGORIZACIÓN DE DESCRIPCIONES
        # =====================================================
        # PRIORITARIAS: se chequean ANTES que los CUITs propios
        # (para evitar que "sircreb Responsable:30711511004" se categorice como transf. propia)
        CATEGORIAS_PRIORITARIAS = [
            ("Retenciones/Percepciones", ["sircreb", "retencion", "retención", "percepcion", "percepción", "recaudacion", "recaudación"]),
            ("Imp. Ley 25.413 Débito",  ["25.413 debito", "25413 debito", "ley 25.413 debito", "ley 25413 debito"]),
            ("Imp. Ley 25.413 Crédito", ["25.413 credito", "25413 credito", "ley 25.413 credito", "ley 25413 credito"]),
        ]

        # GENERALES: se chequean DESPUÉS de los CUITs propios
        CATEGORIAS_GENERALES = [
            ("Comisiones",               ["comision", "comisión"]),
            ("Compras Con Débito",       ["compra con tarjeta de debito", "compra con debito", "compra con deb", "compra debito"]),
            ("Transf. Online Banking",   ["transf. online banking", "transf online banking"]),
            ("Transferencias",           ["transferencia", "pagos ctas propias"]),
            ("Haberes",                  ["haberes", "haber"]),
            ("Impuestos",                ["impuesto", "iibb", "imp."]),
            ("IVA",                      ["iva"]),
            ("Cheques",                  ["cheque"]),
            ("Depósitos",                ["deposito", "depósito"]),
            ("Cajero Automático",        ["cajero", "atm"]),
            ("Débito Automático",        ["débito automático", "debito automatico", "deb.aut"]),
            ("Intereses",                ["interes", "interés"]),
            ("Préstamos",                ["prestamo", "préstamo"]),
            ("Seguros",                  ["seguro"]),
            ("Servicios",                ["servicio"]),
        ]

        def categorizar(descripcion, raw_text=""):
            """Asigna una categoría general a una descripción de movimiento."""
            desc_lower = descripcion.lower()
            texto_busqueda = raw_text if raw_text else descripcion
            texto_sin_espacios = texto_busqueda.replace(" ", "")
            texto_lower = texto_busqueda.lower()

            # PASO 1: Categorías prioritarias (retenciones, sircreb, imp 25413)
            for categoria, keywords in CATEGORIAS_PRIORITARIAS:
                for kw in keywords:
                    if kw in desc_lower or kw in texto_lower:
                        return categoria

            # PASO 2: Transferencias propias por CUIT o Razón Social
            for entry in cuits_propios:
                cuit, razon, label = entry
                if cuit and cuit in texto_sin_espacios:
                    return f"Transf. Propias - {label}"
                if razon and razon.lower() in texto_lower:
                    return f"Transf. Propias - {label}"

            # PASO 3: Categorías generales
            for categoria, keywords in CATEGORIAS_GENERALES:
                for kw in keywords:
                    if kw in desc_lower:
                        return categoria
            return "Otros"

        # --- DEBUG: Mostrar categorizaciones ---
        with st.expander("🔍 DEBUG: Categorizaciones (click para expandir)", expanded=False):
            st.write(f"**CUITs propios configurados:** {cuits_propios}")
            st.write(f"**Total movimientos pesos:** {len(datos_pesos)}")
            for i, mov in enumerate(datos_pesos[:50]):
                fecha, desc, importe, raw = mov
                cat = categorizar(desc, raw)
                cuit_found = ""
                for entry in cuits_propios:
                    c, r, l = entry
                    raw_sin_esp = raw.replace(" ", "")
                    if c and c in raw_sin_esp:
                        cuit_found = f"✅ CUIT '{c}' en raw"
                        break
                    if r and r.lower() in raw.lower():
                        cuit_found = f"✅ Razón '{r}' en raw"
                        break
                if not cuit_found and cuits_propios:
                    cuit_found = "❌ No encontrado"
                emoji = "🟢" if importe > 0 else "🔴"
                st.write(f"{emoji} `{i+1}. [{cat}]` | `{desc[:70]}` | {cuit_found}")

        # =====================================================
        # NUEVA FUNCIÓN: Hojas de Ingresos/Egresos con mini-tablas por categoría
        # =====================================================
        def crear_hoja_agrupada(wb, nombre_hoja, datos, tipo, formato_moneda='"$ "#,##0.00'):
            """
            Crea una hoja con mini-tablas por categoría, todas apiladas verticalmente.
            tipo: 'ingresos' (importe > 0) o 'egresos' (importe < 0)
            """
            ws = wb.create_sheet(title=nombre_hoja)
            ws.sheet_view.showGridLines = False
            # Outline settings: summary row below the detail
            ws.sheet_properties.outlinePr.summaryBelow = True

            if datos and len(datos[0]) == 4:
                df = pd.DataFrame(datos, columns=["Fecha", "Descripcion", "Importe", "RawText"])
            else:
                df = pd.DataFrame(datos, columns=["Fecha", "Descripcion", "Importe"])
                df["RawText"] = ""

            if tipo == "ingresos":
                df_filtrado = df[df["Importe"] > 0].copy()
                fill_header = fill_head_cred
                fill_col = fill_col_cred
                fill_row = fill_row_cred
                titulo_tipo = "INGRESOS"
            else:
                df_filtrado = df[df["Importe"] < 0].copy()
                df_filtrado["Importe"] = df_filtrado["Importe"].abs()
                fill_header = fill_head_deb
                fill_col = fill_col_deb
                fill_row = fill_row_deb
                titulo_tipo = "EGRESOS"

            # Asignar categoría
            if not df_filtrado.empty:
                df_filtrado["Categoria"] = df_filtrado.apply(
                    lambda r: categorizar(r["Descripcion"], r.get("RawText", "")), axis=1
                )

            # Header principal
            ws.merge_cells("A1:C1")
            tit = ws["A1"]
            tit.value = f"SANTANDER {titulo_tipo} ({nombre_hoja.split(' - ')[0]}) - {clean_for_excel(titular_global)}"
            tit.font = Font(size=14, bold=True, color=color_txt_main)
            tit.fill = PatternFill(start_color=color_bg_main, end_color=color_bg_main, fill_type="solid")
            tit.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 25

            # Metadata
            ws["A3"] = "TITULAR"
            ws["A3"].font = Font(bold=True, size=10, color="666666")
            ws.merge_cells("B3:C3")
            ws["B3"] = clean_for_excel(titular_global)
            ws["B3"].alignment = Alignment(horizontal='center')

            ws["A4"] = "PERÍODO"
            ws["A4"].font = Font(bold=True, size=10, color="666666")
            ws.merge_cells("B4:C4")
            ws["B4"] = clean_for_excel(periodo_global)
            ws["B4"].alignment = Alignment(horizontal='center')

            row = 6  # Empezar las mini-tablas

            if df_filtrado.empty:
                ws[f"A{row}"] = "SIN MOVIMIENTOS"
                ws.merge_cells(f"A{row}:C{row}")
                ws[f"A{row}"].alignment = Alignment(horizontal='center')
                ws[f"A{row}"].font = Font(italic=True, color="666666")
            else:
                # Obtener categorías ordenadas alfabéticamente
                categorias_unicas = sorted(df_filtrado["Categoria"].unique())
                total_rows_por_categoria = []  # Para el gran total al final

                for cat in categorias_unicas:
                    df_cat = df_filtrado[df_filtrado["Categoria"] == cat].sort_values("Fecha")

                    # --- Encabezado de categoría ---
                    ws.merge_cells(f"A{row}:C{row}")
                    ws[f"A{row}"] = cat.upper()
                    ws[f"A{row}"].fill = fill_header
                    ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
                    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
                    row += 1

                    # --- Subheaders ---
                    group_start = row  # Inicio del grupo colapsable
                    for col, txt in zip(["A", "B", "C"], ["Fecha", "Descripción", "Importe"]):
                        ws[f"{col}{row}"] = txt
                        ws[f"{col}{row}"].border = thin_border
                        ws[f"{col}{row}"].fill = fill_col
                        ws[f"{col}{row}"].font = Font(bold=True)
                        ws[f"{col}{row}"].alignment = Alignment(horizontal='center')
                    row += 1

                    # --- Datos ---
                    start_data = row
                    for _, r in df_cat.iterrows():
                        ws[f"A{row}"] = r["Fecha"]
                        ws[f"B{row}"] = r["Descripcion"]
                        ws[f"C{row}"] = round(r["Importe"], 2)
                        ws[f"C{row}"].number_format = formato_moneda
                        for c in ["A", "B", "C"]:
                            ws[f"{c}{row}"].border = thin_border
                            ws[f"{c}{row}"].fill = fill_row
                        row += 1

                    group_end = row - 1  # Última fila de datos

                    # --- Agrupar filas (colapsable) ---
                    for r_idx in range(group_start, group_end + 1):
                        ws.row_dimensions[r_idx].outlineLevel = 1
                        ws.row_dimensions[r_idx].hidden = True

                    # --- Total de categoría ---
                    ws.merge_cells(f"A{row}:B{row}")
                    ws[f"A{row}"] = f"TOTAL {cat.upper()}"
                    ws[f"A{row}"].font = Font(bold=True)
                    ws[f"A{row}"].alignment = Alignment(horizontal='right')
                    ws[f"C{row}"] = f"=SUM(C{start_data}:C{row-1})"
                    ws[f"C{row}"].number_format = formato_moneda
                    ws[f"C{row}"].font = Font(bold=True)
                    for c in ["A", "B", "C"]:
                        ws[f"{c}{row}"].border = thin_border
                    total_rows_por_categoria.append(row)
                    row += 2  # Espacio entre tablas

                # === GRAN TOTAL AL FINAL ===
                ws.merge_cells(f"A{row}:B{row}")
                ws[f"A{row}"] = f"GRAN TOTAL {titulo_tipo}"
                ws[f"A{row}"].font = Font(bold=True, size=12)
                ws[f"A{row}"].alignment = Alignment(horizontal='right')
                ws[f"A{row}"].fill = fill_header
                ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
                # Sumar todos los totales de categoría
                formula_parts = "+".join([f"C{tr}" for tr in total_rows_por_categoria])
                ws[f"C{row}"] = f"={formula_parts}"
                ws[f"C{row}"].number_format = formato_moneda
                ws[f"C{row}"].font = Font(bold=True, size=12, color="FFFFFF")
                ws[f"C{row}"].fill = fill_header
                for c in ["A", "B", "C"]:
                    ws[f"{c}{row}"].border = thin_border

            # Anchos
            ws.column_dimensions["A"].width = 15
            ws.column_dimensions["B"].width = 45
            ws.column_dimensions["C"].width = 20

        # =====================================================
        # CREAR HOJAS
        # =====================================================

        # Hoja 1: Pesos (dashboard original)
        crear_hoja_dashboard(wb, "Pesos", datos_pesos, saldo_ini_pesos, saldo_fin_pesos, formato_moneda='"$ "#,##0.00')
        
        # Hoja 2: Pesos - Ingresos
        crear_hoja_agrupada(wb, "Pesos - Ingresos", datos_pesos, "ingresos", formato_moneda='"$ "#,##0.00')
        
        # Hoja 3: Pesos - Egresos
        crear_hoja_agrupada(wb, "Pesos - Egresos", datos_pesos, "egresos", formato_moneda='"$ "#,##0.00')

        # Hojas Dolares (solo si hay datos)
        if datos_dolares or saldo_ini_dolares != 0 or saldo_fin_dolares != 0:
            # Hoja 4: Dolares (dashboard original)
            crear_hoja_dashboard(wb, "Dolares", datos_dolares, saldo_ini_dolares, saldo_fin_dolares, formato_moneda='"U$S "#,##0.00')
            
            # Hoja 5: Dolares - Ingresos
            crear_hoja_agrupada(wb, "Dolares - Ingresos", datos_dolares, "ingresos", formato_moneda='"U$S "#,##0.00')
            
            # Hoja 6: Dolares - Egresos
            crear_hoja_agrupada(wb, "Dolares - Egresos", datos_dolares, "egresos", formato_moneda='"U$S "#,##0.00')

        wb.save(output)
        output.seek(0)
        return output.getvalue()

    except Exception as e:
        import traceback
        st.error(f"Error al procesar el archivo: {str(e)}")
        print(traceback.format_exc())
        return None
