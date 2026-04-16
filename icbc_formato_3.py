import streamlit as st
import io
import pdfplumber
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

# Regex para caracteres ilegales en Excel
ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

def clean_for_excel(text):
    if not text: return ""
    text = str(text)
    text = ILLEGAL_CHARACTERS_RE.sub("", text)
    return text.strip()

def parse_importe(importe_str):
    try:
        clean = importe_str.replace(".", "").replace(",", ".")
        return float(clean)
    except:
        return 0.0

def procesar_icbc_formato_3(archivo_pdf):
    """Procesa ICBC Resumen Mensual de Cuenta Corriente en Pesos.

    Formato Excel: igual al de Santander — dos tablitas lado a lado
    (Créditos | Débitos) con control de saldos.

    Fixes de extracción:
    1. Detecta el signo '-' después del importe para Débito vs Crédito.
    2. Cuando hay 2 importes en la línea, el primero es la transacción y
       el segundo es el saldo acumulado de fin de hoja (se descarta).
    3. Detecta el número de cuenta a partir de 'CUENTA CORRIENTE'.
    4. El regex de período acepta DD-MM-YYYY y DD/MM/YYYY.
    5. Extrae saldo inicial (SALDO ULTIMO EXTRACTO) y saldo final.
    """
    st.info("Procesando archivo ICBC (Resumen Mensual Cuenta Corriente)...")

    try:
        texto_completo = ""
        with pdfplumber.open(io.BytesIO(archivo_pdf.read())) as pdf:
            # El PDF de ICBC suele contener el extracto duplicado (anverso + reverso
            # encuadernados o una segunda copia de cortesía). Detectamos cuándo
            # aparece por segunda vez "SALDO ULTIMO EXTRACTO" para cortar ahí y
            # procesar solo la primera copia, evitando duplicar todos los importes.
            saldo_ini_visto = False
            for page in pdf.pages:
                t = page.extract_text()
                if not t:
                    continue
                if "SALDO ULTIMO EXTRACTO" in t:
                    if saldo_ini_visto:
                        break   # Segunda copia — detenemos la lectura
                    saldo_ini_visto = True
                texto_completo += t + "\n"

        lineas = texto_completo.splitlines()

        # ── Metadata ─────────────────────────────────────────────────────────
        titular_global = "Sin Especificar"
        periodo_global = "Sin Especificar"
        anio_global = "2026"
        saldo_ini = 0.0
        saldo_fin = 0.0

        # Titular: primera línea no vacía antes de los encabezados técnicos
        for l in lineas[:5]:
            cleaned = l.strip()
            if cleaned and "PERIODO" not in l.replace(" ", "").upper():
                titular_global = cleaned
                break

        # Período: "PERIODO DD-MM-YYYY AL DD-MM-YYYY"
        for l in lineas[:40]:
            clean_l = l.replace(" ", "")
            if "PERIODO" in clean_l.upper():
                fechas = re.findall(r"(\d{2}[-/]\d{2}[-/]\d{4})", clean_l)
                if len(fechas) >= 2:
                    f_ini = fechas[0].replace("-", "/")
                    f_fin = fechas[1].replace("-", "/")
                    periodo_global = f"Del {f_ini} al {f_fin}"
                    anio_global = re.split(r"[-/]", fechas[0])[-1]
                    break

        # Saldo inicial: "SALDO ULTIMO EXTRACTO AL 28/02/2026 599.813.992,65"
        regex_saldo_ini = re.compile(r"SALDO ULTIMO EXTRACTO.*?([\d\.]+,\d{2})")
        # Saldo final: "SALDO FINAL AL 31/03/2026 138.937.027,92"
        regex_saldo_fin = re.compile(r"SALDO FINAL AL.*?([\d\.]+,\d{2})")

        for l in lineas:
            m = regex_saldo_ini.search(l)
            if m and saldo_ini == 0.0:
                saldo_ini = parse_importe(m.group(1))
            m = regex_saldo_fin.search(l)
            if m:
                saldo_fin = parse_importe(m.group(1))

        # ── Regex de trabajo ─────────────────────────────────────────────────
        regex_linea = re.compile(r"^(\d{2}-\d{2})\s+(.+)$")
        # Captura el número con formato miles + signo opcional al final
        regex_importe_signo = re.compile(r"(\d{1,3}(?:\.\d{3})*,\d{2})(-?)")
        regex_cuenta = re.compile(
            r"CUENTA CORRIENTE.*?N[°oº]?\s*([\w/]+)", re.IGNORECASE
        )

        SALTAR = (
            "FECHA", "HOJA N", "SALDO HOJA", "SALDO PAGINA",
            "CONTINUA", "TOT.IMP", "TOTAL RECAUD", "F.EXTDOR",
            "F.EXTSTD", "SALDO FINAL", "SALDO ULTIMO",
        )

        # ── Parseo de movimientos ─────────────────────────────────────────────
        movimientos = []
        cuenta_actual = "Principal"

        for l in lineas:
            l_s = l.strip()

            m_cta = regex_cuenta.search(l_s)
            if m_cta:
                cuenta_actual = m_cta.group(1)
                continue

            if any(k in l_s for k in SALTAR):
                continue

            match = regex_linea.match(l_s)
            if not match:
                continue

            fecha_dia_mes = match.group(1)
            resto = match.group(2)

            matches_imp = list(regex_importe_signo.finditer(resto))
            if not matches_imp:
                continue

            # Primer importe = transacción; si hay 2+, el último es el saldo
            m_tx = matches_imp[0]
            monto_str = m_tx.group(1)
            es_debito = m_tx.group(2) == "-"
            importe = parse_importe(monto_str)

            descripcion = resto[: m_tx.start()].strip()
            descripcion = re.sub(r"\s+\d{2}-\d{2}\s+", " ", descripcion).strip()

            fecha = f"{fecha_dia_mes.replace('-', '/')}/{anio_global}"

            movimientos.append({
                "Fecha": fecha,
                "Cuenta": cuenta_actual,
                "Descripcion": clean_for_excel(descripcion),
                "Importe": importe,
                "Es_Debito": es_debito,
            })

        if not movimientos:
            st.error("No se encontraron movimientos en este archivo.")
            return None

        df = pd.DataFrame(movimientos)
        creditos = df[~df["Es_Debito"]].copy()
        debitos  = df[ df["Es_Debito"]].copy()

        # ── Excel ─────────────────────────────────────────────────────────────
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Movimientos"
        ws.sheet_view.showGridLines = False

        # ── Paleta de colores (igual que Santander) ───────────────────────────
        color_rojo_icbc = "C5001A"
        color_blanco = "FFFFFF"
        fmt_moneda = '"$ "#,##0.00'

        thin_border = Border(
            left=Side(style="thin", color="A6A6A6"),
            right=Side(style="thin", color="A6A6A6"),
            top=Side(style="thin", color="A6A6A6"),
            bottom=Side(style="thin", color="A6A6A6"),
        )

        fill_head_cred = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        fill_col_cred  = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
        fill_row_cred  = PatternFill(start_color="F2F9F1", end_color="F2F9F1", fill_type="solid")

        fill_head_deb  = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        fill_col_deb   = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
        fill_row_deb   = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")

        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_font = Font(color="9C0006", bold=True)

        # ── Fila 1: Título principal ──────────────────────────────────────────
        ws.merge_cells("A1:G1")
        tit = ws["A1"]
        tit.value = f"REPORTE ICBC — {clean_for_excel(titular_global)}"
        tit.font = Font(size=14, bold=True, color=color_blanco)
        tit.fill = PatternFill(start_color=color_rojo_icbc, end_color=color_rojo_icbc, fill_type="solid")
        tit.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # ── Filas 3-4: Metadata ───────────────────────────────────────────────
        ws["A3"] = "SALDO INICIAL"
        ws["A3"].font = Font(bold=True, size=10, color="666666")
        ws["B3"] = saldo_ini
        ws["B3"].number_format = fmt_moneda
        ws["B3"].font = Font(bold=True, size=11)
        ws["B3"].border = Border(bottom=Side(style="thin", color="DDDDDD"))

        ws["A4"] = "SALDO FINAL"
        ws["A4"].font = Font(bold=True, size=10, color="666666")
        ws["B4"] = saldo_fin
        ws["B4"].number_format = fmt_moneda
        ws["B4"].font = Font(bold=True, size=11)
        ws["B4"].border = Border(bottom=Side(style="thin", color="DDDDDD"))

        ws["D3"] = "TITULAR"
        ws["D3"].font = Font(bold=True, size=10, color="666666")
        ws.merge_cells("E3:G3")
        ws["E3"] = clean_for_excel(titular_global)
        ws["E3"].alignment = Alignment(horizontal="center")
        ws["E3"].font = Font(bold=True)

        ws["D4"] = "PERÍODO"
        ws["D4"].font = Font(bold=True, size=10, color="666666")
        ws.merge_cells("E4:G4")
        ws["E4"] = clean_for_excel(periodo_global)
        ws["E4"].alignment = Alignment(horizontal="center")

        # ── Fila 6-7: Control de saldos ───────────────────────────────────────
        ws["D6"] = "CONTROL DE SALDOS"
        ws["D6"].font = Font(bold=True, size=10, color="666666")
        # Se completa con fórmula después de construir las tablas
        ws["E6"].border = thin_border
        ws["E6"].font = Font(bold=True, size=12)
        ws.conditional_formatting.add(
            "E6",
            CellIsRule(operator="notEqual", formula=["0"], stopIfTrue=True,
                       fill=red_fill, font=red_font)
        )

        # ── Fila 10+: Tablas ──────────────────────────────────────────────────
        F_HEADER = 10

        # Cabeceras de sección
        ws.merge_cells(f"A{F_HEADER}:C{F_HEADER}")
        ws[f"A{F_HEADER}"] = "CRÉDITOS"
        ws[f"A{F_HEADER}"].fill = fill_head_cred
        ws[f"A{F_HEADER}"].font = Font(bold=True, color=color_blanco)
        ws[f"A{F_HEADER}"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(f"E{F_HEADER}:G{F_HEADER}")
        ws[f"E{F_HEADER}"] = "DÉBITOS"
        ws[f"E{F_HEADER}"].fill = fill_head_deb
        ws[f"E{F_HEADER}"].font = Font(bold=True, color=color_blanco)
        ws[f"E{F_HEADER}"].alignment = Alignment(horizontal="center", vertical="center")

        # Sub-encabezados de columna
        for col, txt in zip(
            ["A", "B", "C", "E", "F", "G"],
            ["Fecha", "Descripción", "Importe", "Fecha", "Descripción", "Importe"]
        ):
            c = ws[f"{col}{F_HEADER + 1}"]
            c.value = txt
            c.border = thin_border
            c.alignment = Alignment(horizontal="center")
            c.fill = fill_col_cred if col in ("A", "B", "C") else fill_col_deb
            c.font = Font(bold=True)

        # ── Llenar tabla CRÉDITOS ─────────────────────────────────────────────
        row = F_HEADER + 2
        start_cred = row

        if creditos.empty:
            ws[f"A{row}"] = "SIN MOVIMIENTOS"
            ws.merge_cells(f"A{row}:C{row}")
            ws[f"A{row}"].alignment = Alignment(horizontal="center")
            ws[f"A{row}"].font = Font(italic=True, color="666666")
            row += 1
        else:
            for _, r in creditos.iterrows():
                ws[f"A{row}"] = r["Fecha"]
                ws[f"B{row}"] = r["Descripcion"]
                ws[f"C{row}"] = r["Importe"]
                ws[f"C{row}"].number_format = fmt_moneda
                for c in ("A", "B", "C"):
                    ws[f"{c}{row}"].border = thin_border
                    ws[f"{c}{row}"].fill = fill_row_cred
                row += 1

        total_cred_row = row
        ws.merge_cells(f"A{total_cred_row}:B{total_cred_row}")
        ws[f"A{total_cred_row}"] = "TOTAL CRÉDITOS"
        ws[f"A{total_cred_row}"].font = Font(bold=True)
        ws[f"A{total_cred_row}"].alignment = Alignment(horizontal="right")
        ws[f"C{total_cred_row}"] = f"=SUM(C{start_cred}:C{total_cred_row - 1})"
        ws[f"C{total_cred_row}"].number_format = fmt_moneda
        ws[f"C{total_cred_row}"].font = Font(bold=True)
        for c in ("A", "B", "C"):
            ws[f"{c}{total_cred_row}"].border = thin_border

        # ── Llenar tabla DÉBITOS ──────────────────────────────────────────────
        row = F_HEADER + 2
        start_deb = row

        if debitos.empty:
            ws[f"E{row}"] = "SIN MOVIMIENTOS"
            ws.merge_cells(f"E{row}:G{row}")
            ws[f"E{row}"].alignment = Alignment(horizontal="center")
            ws[f"E{row}"].font = Font(italic=True, color="666666")
            row += 1
        else:
            for _, r in debitos.iterrows():
                ws[f"E{row}"] = r["Fecha"]
                ws[f"F{row}"] = r["Descripcion"]
                ws[f"G{row}"] = r["Importe"]
                ws[f"G{row}"].number_format = fmt_moneda
                for c in ("E", "F", "G"):
                    ws[f"{c}{row}"].border = thin_border
                    ws[f"{c}{row}"].fill = fill_row_deb
                row += 1

        total_deb_row = row
        ws.merge_cells(f"E{total_deb_row}:F{total_deb_row}")
        ws[f"E{total_deb_row}"] = "TOTAL DÉBITOS"
        ws[f"E{total_deb_row}"].font = Font(bold=True)
        ws[f"E{total_deb_row}"].alignment = Alignment(horizontal="right")
        ws[f"G{total_deb_row}"] = f"=SUM(G{start_deb}:G{total_deb_row - 1})"
        ws[f"G{total_deb_row}"].number_format = fmt_moneda
        ws[f"G{total_deb_row}"].font = Font(bold=True)
        for c in ("E", "F", "G"):
            ws[f"{c}{total_deb_row}"].border = thin_border

        # ── Fórmula de control: Saldo Ini + Créditos - Débitos = Saldo Fin ────
        # Si da 0 → todo cuadra; si no → resalta en rojo
        ws["E6"] = (
            f"=ROUND(B3 + C{total_cred_row} - G{total_deb_row} - B4, 2)"
        )
        ws["E6"].number_format = fmt_moneda

        # ── Anchos de columna ─────────────────────────────────────────────────
        ws.column_dimensions["A"].width = 13
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 13
        ws.column_dimensions["F"].width = 45
        ws.column_dimensions["G"].width = 18

        wb.save(output)
        output.seek(0)
        return output.getvalue()

    except Exception as e:
        import traceback
        st.error(f"Error procesando ICBC Formato 3: {e}")
        print(traceback.format_exc())
        return None
