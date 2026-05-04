import streamlit as st
import PyPDF2
import re
import pandas as pd
import io
import sys


def limpiar_nombre_hoja(nombre):
    """Limpia el nombre para que sea válido como nombre de hoja de Excel"""
    # Caracteres no permitidos en nombres de hojas de Excel
    caracteres_invalidos = ["\\", "/", "*", "[", "]", ":", "?"]
    nombre_limpio = nombre
    for char in caracteres_invalidos:
        nombre_limpio = nombre_limpio.replace(char, "_")

    # Limitar longitud a 31 caracteres (límite de Excel)
    if len(nombre_limpio) > 31:
        nombre_limpio = nombre_limpio[:31]

    return nombre_limpio


def procesar_mercadopago(archivo_pdf):
    """Procesa archivos PDF de MercadoPago"""

    try:
        # Reinicializar el archivo para lectura
        archivo_pdf.seek(0)

        # Abrir y leer el archivo PDF
        reader = PyPDF2.PdfReader(io.BytesIO(archivo_pdf.read()))
        texto = "".join(page.extract_text() + "\n" for page in reader.pages)

        lineas = texto.splitlines()

        print(
            f"[MercadoPago] file={getattr(archivo_pdf, 'name', '?')} "
            f"size={getattr(archivo_pdf, 'size', '?')} "
            f"pypdf2={PyPDF2.__version__} pages={len(reader.pages)} "
            f"text_len={len(texto)} "
            f"has_saldo_inicial={'saldo inicial' in texto.lower()} "
            f"has_saldo_final={'saldo final' in texto.lower()}",
            file=sys.stderr, flush=True,
        )

        # Variables para corte de página (descripción huérfana)
        prefijo_pendiente = ""
        just_completed_movement = False

        # Variables para almacenar la información
        saldo_inicial = None
        saldo_final = None
        movimientos = []
        nombre_titular = None
        cvu = None
        periodo = None

        # Detección de saldos sobre el texto completo (tolerante a saltos de
        # línea o espaciado raro de PyPDF2 entre containers de Streamlit Cloud).
        m_ini = re.search(r"Saldo\s+inicial\s*:\s*\$?\s*([\d.,]+)", texto, re.IGNORECASE)
        if m_ini:
            saldo_inicial = m_ini.group(1)
        m_fin = re.search(r"Saldo\s+final\s*:\s*\$?\s*([\d.,]+)", texto, re.IGNORECASE)
        if m_fin:
            saldo_final = m_fin.group(1)

        # Procesar líneas
        i = 0
        while i < len(lineas):
            linea = lineas[i].strip()
            
            # Ignorar líneas vacías
            if not linea:
                i += 1
                continue

            # Ignorar líneas vacías después de la limpieza
            if not linea:
                i += 1
                continue

            # Extraer nombre del titular (primera línea después de "RESUMEN DE CUENTA")
            # La línea anterior puede tener basura como "1/61RESUMEN DE CUENTA"
            if i > 0 and "RESUMEN DE CUENTA" in lineas[i - 1] and linea:
                # Nos aseguramos de no sobreescribir si ya lo tenemos (por si aparece otra vez más adelante)
                if not nombre_titular:
                    nombre_titular = linea

            # Extraer Período (Formato: "Del 1 al ... Periodo:")
            if "Periodo:" in linea or "Período:" in linea:
                 # Capturar lo que está ANTES de "Periodo:"
                 match_periodo = re.search(r"(.*)(Periodo:|Período:)", linea, re.IGNORECASE)
                 if match_periodo:
                     periodo = match_periodo.group(1).strip()

            # Extraer CVU
            if linea.startswith("CVU:"):
                cvu_match = re.search(r"CVU:\s*(\d+)", linea)
                if cvu_match:
                    cvu = cvu_match.group(1)

            match_fecha_inicio = re.search(r"(\d{2}-\d{2}-\d{4})", linea)
            
            # Solo consideramos que es inicio de movimiento si la fecha aparece al principio (primeros 20 chars)
            if match_fecha_inicio and match_fecha_inicio.start() < 20:
                # Limpiar basura anterior a la fecha
                linea = linea[match_fecha_inicio.start():]
                # Prefijar texto huérfano de corte de página
                if prefijo_pendiente:
                    fecha_str = linea[:10]
                    resto = linea[10:].strip()
                    linea = fecha_str + " " + prefijo_pendiente + " " + resto
                    prefijo_pendiente = ""
                just_completed_movement = False
                linea_movimiento = linea

                # Verificar si la línea actual contiene los montos
                lineas_extra = 0
                while not re.search(r"\$\s*-?[\d,]+\.?\d*", linea_movimiento) and lineas_extra < 20:
                    if i + 1 < len(lineas):
                        linea_siguiente_check = lineas[i + 1].strip()
                        
                        # Detectar y saltar encabezados de página
                        if re.match(r"^\d+/\d+\s*Fecha", linea_siguiente_check):
                            i += 1
                            continue
                        
                        if re.match(r"^\d+\s*/\s*\d+$", linea_siguiente_check):
                            i += 1
                            continue

                        # Verificar inicio nuevo movimiento
                        match_fecha_next = re.search(r"(\d{2}-\d{2}-\d{4})", linea_siguiente_check)
                        if match_fecha_next and match_fecha_next.start() < 20:
                            break 
                            
                        linea_movimiento += " " + linea_siguiente_check
                        i += 1
                        lineas_extra += 1
                    else:
                        break
                
                # Limpiar la línea: quitar saltos de línea internos y espacios extra
                linea_movimiento = " ".join(linea_movimiento.split())

                # Extraer fecha usando regex (primeros 10 caracteres en formato DD-MM-YYYY)
                fecha_match = re.match(r"^(\d{2}-\d{2}-\d{4})", linea_movimiento)
                if fecha_match:
                    fecha = fecha_match.group(1)

                    # Buscar montos con regex mejorado - incluir decimales opcionales
                    # Primero buscar todos los fragmentos de números
                    fragmentos_numericos = re.findall(r"[\d.,]+", linea_movimiento)

                    # Reconstruir montos válidos
                    montos_validos = []
                    i_frag = 0
                    while i_frag < len(fragmentos_numericos):
                        fragmento = fragmentos_numericos[i_frag]

                        # Si el fragmento termina en coma, buscar el siguiente fragmento como decimales
                        if fragmento.endswith(",") and i_frag + 1 < len(
                            fragmentos_numericos
                        ):
                            siguiente = fragmentos_numericos[i_frag + 1]
                            # Si el siguiente fragmento son solo 2 dígitos, es parte decimal
                            if re.match(r"^\d{2}$", siguiente):
                                monto_completo = fragmento + siguiente
                                montos_validos.append(monto_completo)
                                i_frag += 2  # Saltar el siguiente fragmento
                                continue

                        # Verificar si es un monto válido (formato argentino)
                        if re.match(r"^\d{1,3}(?:\.\d{3})*(?:,\d{2})?$", fragmento):
                            montos_validos.append(fragmento)

                        i_frag += 1

                    if len(montos_validos) >= 2:
                        importe = montos_validos[-2]  # Penúltimo monto válido

                        # Detectar si el importe es negativo buscando el signo - antes del monto
                        # Buscar la posición del importe en la línea
                        posicion_importe = linea_movimiento.find(importe)
                        if posicion_importe > 0:
                            # Revisar los caracteres antes del importe para buscar el signo -
                            texto_antes = linea_movimiento[:posicion_importe]
                            # Buscar el último $ seguido opcionalmente de espacios y -
                            if re.search(r"\$\s*-\s*$", texto_antes):
                                importe = "-" + importe

                        # Extraer descripción (todo después de la fecha hasta antes del ID y montos)
                        # Remover la fecha del inicio
                        resto_linea = linea_movimiento[10:].strip()  # Quitar los primeros 10 caracteres (fecha)

                        # Extraer descripción (todo después de la fecha hasta antes del ID y montos)
                        # Remover la fecha del inicio
                        resto_linea = linea_movimiento[10:].strip()  # Quitar los primeros 10 caracteres (fecha)

                        # Buscar el ID (número largo de 10 o más dígitos)
                        match_id = re.search(r"(\d{10,})", resto_linea)
                        
                        if match_id:
                            # Cortamos todo lo que está ANTES de ese ID
                            inicio_id = match_id.start()
                            descripcion = resto_linea[:inicio_id].strip()
                            # Limpieza extra: a veces quedan comas al final
                            if descripcion.endswith(","):
                                descripcion = descripcion[:-1].strip()
                        else:
                            # Fallback...
                            if importe:
                                importe_clean = importe.replace("-", "").strip()
                                idx_importe = resto_linea.find(importe_clean)
                                if idx_importe != -1: 
                                    sub = resto_linea[:idx_importe]
                                    if sub.strip().endswith("$"):
                                        sub = sub.strip()[:-1]
                                    descripcion = sub.strip()
                                else:
                                    descripcion = resto_linea
                            else:
                                descripcion = resto_linea

                        movimiento = {
                            "Fecha": fecha,
                            "Descripcion": descripcion,
                            "Importe": importe,
                        }
                        movimientos.append(movimiento)
                        just_completed_movement = True

                    elif len(montos_validos) == 1:
                        # Si solo hay un monto válido
                        importe = montos_validos[0]

                        # Detectar si el importe es negativo
                        posicion_importe = linea_movimiento.find(importe)
                        if posicion_importe > 0:
                            texto_antes = linea_movimiento[:posicion_importe]
                            if re.search(r"\$\s*-\s*$", texto_antes):
                                importe = "-" + importe

                        # Extraer descripción
                        resto_linea = linea_movimiento[10:].strip()  # Quitar los primeros 10 caracteres (fecha)

                        # Intento 1: Buscar ID de 10+ dígitos
                        match_id = re.search(r"(\d{10,})", resto_linea)
                        
                        if match_id:
                            # Cortamos todo lo que está ANTES de ese ID
                            inicio_id = match_id.start()
                            descripcion = resto_linea[:inicio_id].strip()
                            # Limpieza extra
                            if descripcion.endswith(","):
                                descripcion = descripcion[:-1].strip()
                        else:
                            # Fallback: usar el importe como corte
                            if importe:
                                importe_clean = importe.replace("-", "").strip()
                                idx_importe = resto_linea.find(importe_clean)
                                if idx_importe != -1:
                                    sub = resto_linea[:idx_importe]
                                    if sub.strip().endswith("$"):
                                        sub = sub.strip()[:-1]
                                    descripcion = sub.strip()
                                else:
                                    descripcion = resto_linea
                            else:
                                descripcion = resto_linea

                        movimiento = {
                            "Fecha": fecha,
                            "Descripcion": descripcion,
                            "Importe": importe,
                        }
                        movimientos.append(movimiento)
                        just_completed_movement = True

            else:
                # Detectar texto descriptivo huérfano por corte de página
                # Solo si acabamos de completar un movimiento
                if just_completed_movement and linea:
                    es_ignorable = (
                        re.match(r"^\d+\s*/\s*\d+", linea) or
                        re.match(r"^Fecha\s+Descripci", linea) or
                        re.match(r"^operaci", linea, re.IGNORECASE) or
                        "$" in linea or
                        "Saldo inicial:" in linea or
                        "Saldo final:" in linea or
                        "RESUMEN DE CUENTA" in linea or
                        "Periodo:" in linea or "Período:" in linea or
                        linea.startswith("CVU:") or
                        "mercadopago" in linea.lower() or
                        "Mercado Libre" in linea or
                        re.match(r"^\d+$", linea)
                    )
                    if not es_ignorable:
                        prefijo_pendiente = (prefijo_pendiente + " " + linea).strip() if prefijo_pendiente else linea
                    else:
                        just_completed_movement = False

            i += 1

        # Crear el archivo Excel
        if saldo_inicial and saldo_final:
            try:
                output = io.BytesIO()

                # Crear DataFrame con los movimientos
                df = pd.DataFrame(movimientos)

                # Función simple para convertir importes a numérico
                def convertir_a_numerico(importe_str):
                    """Convierte importe a numérico tratando punto como separador de miles y coma como decimal"""
                    if not importe_str:
                        return 0.0

                    # Limpiar espacios y detectar signo
                    importe_str = str(importe_str).strip()
                    signo = -1 if importe_str.startswith("-") else 1
                    importe_str = importe_str.lstrip("-").strip()

                    # Formato argentino: punto = separador de miles, coma = decimal
                    # Ejemplos: -1.400 = -1400, 33.688,50 = 33688.50, 1.234,56 = 1234.56

                    if "," in importe_str:
                        # Tiene decimales: 33.688,50
                        partes = importe_str.split(",")
                        parte_entera = partes[0].replace(
                            ".", ""
                        )  # Quitar puntos de miles: 33.688 -> 33688
                        parte_decimal = partes[1]  # Mantener decimales: 50
                        numero_str = f"{parte_entera}.{parte_decimal}"  # 33688.50
                    else:
                        # Solo enteros con separador de miles: 1.400 -> 1400
                        numero_str = importe_str.replace(
                            ".", ""
                        )  # Quitar puntos: 1.400 -> 1400

                    try:
                        return signo * float(numero_str)
                    except ValueError:
                        # Si no se puede convertir, devolver 0
                        return 0.0

                # Convertir la columna Importe a numérico
                if not df.empty:
                    df["Importe"] = df["Importe"].apply(convertir_a_numerico)

                # Separar movimientos en créditos y débitos
                creditos = (
                    df[df["Importe"] > 0].copy()
                    if not df.empty
                    else pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])
                )
                debitos = (
                    df[df["Importe"] < 0].copy()
                    if not df.empty
                    else pd.DataFrame(columns=["Fecha", "Descripcion", "Importe"])
                )

                # Convertir débitos a valores absolutos para mejor visualización
                if not debitos.empty:
                    debitos["Importe"] = debitos["Importe"].abs()

                # Nombre de la hoja
                if cvu:
                    # Usar el CVU como nombre de la hoja
                    nombre_hoja = str(cvu)
                else:
                    # Fallback si no hay CVU
                    nombre_hoja = (
                        f"MercadoPago {nombre_titular[:15] if nombre_titular else 'Cuenta'}"
                    )
                nombre_limpio = limpiar_nombre_hoja(nombre_hoja)

                # Crear el workbook y worksheet manualmente
                from openpyxl import Workbook
                from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
                from openpyxl.formatting.rule import CellIsRule

                wb = Workbook()
                ws = wb.active
                
                # ---------------------------------------------------------
                # 1. PREPARACIÓN VISUAL GENERAL (Estilo Dashboard)
                # ---------------------------------------------------------
                ws.title = nombre_limpio
                ws.sheet_view.showGridLines = False  # Ocultar líneas de cuadrícula

                # Definición de Estilos y Colores
                # Bordes
                thin_border = Border(left=Side(style='thin', color="A6A6A6"), 
                                     right=Side(style='thin', color="A6A6A6"), 
                                     top=Side(style='thin', color="A6A6A6"), 
                                     bottom=Side(style='thin', color="A6A6A6"))
                
                # Colores Corporativos / Semánticos
                color_bg_main = "2C3E50" # Azul noche (Título principal)
                color_txt_main = "FFFFFF"
                
                # Débitos (Rojos)
                fill_head_deb = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
                fill_col_deb = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid") 
                fill_row_deb = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid") # Salmón muy suave

                # Créditos (Verdes)
                fill_head_cred = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
                fill_col_cred = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
                fill_row_cred = PatternFill(start_color="F2F9F1", end_color="F2F9F1", fill_type="solid") # Verde muy suave

                # ---------------------------------------------------------
                # 2. ENCABEZADO DEL REPORTE
                # ---------------------------------------------------------
                ws.merge_cells("A1:G1")
                titulo_main = ws["A1"]
                titulo_main.value = f"REPORTE DE MOVIMIENTOS - {nombre_hoja}"
                titulo_main.font = Font(size=14, bold=True, color=color_txt_main)
                titulo_main.fill = PatternFill(start_color=color_bg_main, end_color=color_bg_main, fill_type="solid")
                titulo_main.alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[1].height = 25

                # ---------------------------------------------------------
                # 3. RESUMEN DE SALDOS (Caja estilo tarjeta)
                # ---------------------------------------------------------
                # Fila 3 y 4 para saldos
                ws["A3"] = "SALDO INICIAL"
                ws["A3"].font = Font(bold=True, size=10, color="666666")
                
                if saldo_inicial:
                    ws["B3"] = convertir_a_numerico(saldo_inicial)
                    ws["B3"].number_format = '"$ "#,##0.00'
                    ws["B3"].font = Font(bold=True, size=11)
                else:
                    ws["B3"] = 0

                ws["A4"] = "SALDO FINAL"
                ws["A4"].font = Font(bold=True, size=10, color="666666")

                if saldo_final:
                    ws["B4"] = convertir_a_numerico(saldo_final)
                    ws["B4"].number_format = '"$ "#,##0.00'
                    ws["B4"].font = Font(bold=True, size=11)
                else:
                    ws["B4"] = 0
                
                # Borde discreto para los saldos
                for r in [3, 4]:
                    ws[f"B{r}"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

                # ---------------------------------------------------------
                # 3.1. INFORMACIÓN ADICIONAL (Titular / Período) - Centro derecha
                # ---------------------------------------------------------
                ws["D3"] = "TITULAR"
                ws["D3"].font = Font(bold=True, size=10, color="666666")
                ws["D3"].alignment = Alignment(horizontal='right')
                
                ws["E3"] = nombre_titular if nombre_titular else "Desconocido"
                ws["E3"].font = Font(bold=True, size=11)
                ws["E3"].alignment = Alignment(horizontal='center')
                ws.merge_cells("E3:G3")
                
                ws["D4"] = "PERÍODO"
                ws["D4"].font = Font(bold=True, size=10, color="666666")
                ws["D4"].alignment = Alignment(horizontal='right')

                ws["E4"] = periodo if periodo else "Desconocido"
                ws["E4"].font = Font(bold=True, size=11)
                ws["E4"].alignment = Alignment(horizontal='center')
                ws.merge_cells("E4:G4")

                # Bordes para info
                for r in [3, 4]:
                     for c in ["E", "F", "G"]:
                        ws[f"{c}{r}"].border = Border(bottom=Side(style='thin', color="DDDDDD"))

                # ---------------------------------------------------------
                # 3.2. CONTROL DE INTEGRIDAD (Fila 7, Separado -> D6/D7 Centrado)
                # ---------------------------------------------------------
                # Lo ponemos vertical en columna D
                ws["D6"] = "CONTROL DE SALDOS"
                ws["D6"].font = Font(bold=True, size=10, color="666666")
                ws["D6"].alignment = Alignment(horizontal='center', vertical='bottom')

                # Reservamos D7 para el valor
                cell_control = ws["D7"]
                cell_control.font = Font(bold=True, size=12)
                cell_control.alignment = Alignment(horizontal='center', vertical='center')
                cell_control.border = Border(bottom=Side(style='thin', color="A6A6A6"), 
                                             top=Side(style='thin', color="A6A6A6"),
                                             left=Side(style='thin', color="A6A6A6"),
                                             right=Side(style='thin', color="A6A6A6"))

                # ---------------------------------------------------------
                # 4. TABLAS DE DATOS
                # ---------------------------------------------------------
                fila_inicio_tablas = 10
                
                # HEADERS FIJOS
                # CRÉDITOS (A-C)
                f_header = fila_inicio_tablas
                ws.merge_cells(f"A{f_header}:C{f_header}")
                ws[f"A{f_header}"] = "CRÉDITOS" 
                ws[f"A{f_header}"].fill = fill_head_cred
                ws[f"A{f_header}"].font = Font(bold=True, color="FFFFFF")
                ws[f"A{f_header}"].alignment = Alignment(horizontal='center')
                ws[f"A{f_header}"].border = thin_border
                
                headers = ["Fecha", "Descripción", "Importe"]
                cols_cred = ["A", "B", "C"]
                f_sub = f_header + 1
                for i, h in enumerate(headers):
                    c = ws[f"{cols_cred[i]}{f_sub}"]
                    c.value = h
                    c.fill = fill_col_cred
                    c.font = Font(bold=True)
                    c.alignment = Alignment(horizontal='center')
                    c.border = thin_border
                
                # DÉBITOS (E-G)
                ws.merge_cells(f"E{f_header}:G{f_header}")
                ws[f"E{f_header}"] = "DÉBITOS" 
                ws[f"E{f_header}"].fill = fill_head_deb
                ws[f"E{f_header}"].font = Font(bold=True, color="FFFFFF")
                ws[f"E{f_header}"].alignment = Alignment(horizontal='center')
                ws[f"E{f_header}"].border = thin_border
                
                cols_deb = ["E", "F", "G"]
                for i, h in enumerate(headers):
                    c = ws[f"{cols_deb[i]}{f_sub}"]
                    c.value = h
                    c.fill = fill_col_deb
                    c.font = Font(bold=True)
                    c.alignment = Alignment(horizontal='center')
                    c.border = thin_border
                    
                # --- LLENADO DE DATOS (PARALELO) ---
                fila_dato_start = f_sub + 1
                
                # 1. CRÉDITOS
                f_cred = fila_dato_start
                if creditos.empty:
                    ws.merge_cells(f"A{f_cred}:C{f_cred}")
                    ws[f"A{f_cred}"] = "SIN MOVIMIENTOS"
                    ws[f"A{f_cred}"].font = Font(italic=True, color="666666")
                    ws[f"A{f_cred}"].alignment = Alignment(horizontal='center')
                    ws[f"A{f_cred}"].border = thin_border
                    f_cred += 1
                else:
                    start_c = f_cred
                    for _, r in creditos.iterrows():
                        ws[f"A{f_cred}"] = r["Fecha"]  # MP dates are strings usually
                        ws[f"A{f_cred}"].fill = fill_row_cred
                        ws[f"A{f_cred}"].alignment = Alignment(horizontal='center')
                        ws[f"A{f_cred}"].border = thin_border

                        ws[f"B{f_cred}"] = str(r["Descripcion"])
                        ws[f"B{f_cred}"].fill = fill_row_cred
                        ws[f"B{f_cred}"].border = thin_border

                        ws[f"C{f_cred}"] = r["Importe"]
                        ws[f"C{f_cred}"].number_format = '"$ "#,##0.00'
                        ws[f"C{f_cred}"].fill = fill_row_cred
                        ws[f"C{f_cred}"].border = thin_border
                        f_cred += 1
                    
                    # Total Créditos
                    ws.merge_cells(f"A{f_cred}:B{f_cred}")
                    ws[f"A{f_cred}"] = "TOTAL CRÉDITOS"
                    ws[f"A{f_cred}"].font = Font(bold=True)
                    ws[f"A{f_cred}"].alignment = Alignment(horizontal='right')
                    ws[f"A{f_cred}"].fill = fill_col_cred
                    ws[f"A{f_cred}"].border = thin_border
                    
                    ws[f"C{f_cred}"] = f"=SUM(C{start_c}:C{f_cred-1})"
                    ws[f"C{f_cred}"].number_format = '"$ "#,##0.00'
                    ws[f"C{f_cred}"].font = Font(bold=True)
                    ws[f"C{f_cred}"].fill = fill_col_cred
                    ws[f"C{f_cred}"].border = thin_border
                    f_cred += 1

                # 2. DÉBITOS
                f_deb = fila_dato_start
                if debitos.empty:
                    ws.merge_cells(f"E{f_deb}:G{f_deb}")
                    ws[f"E{f_deb}"] = "SIN MOVIMIENTOS"
                    ws[f"E{f_deb}"].font = Font(italic=True, color="666666")
                    ws[f"E{f_deb}"].alignment = Alignment(horizontal='center')
                    ws[f"E{f_deb}"].border = thin_border
                    f_deb += 1
                else:
                    start_d = f_deb
                    for _, r in debitos.iterrows():
                        ws[f"E{f_deb}"] = r["Fecha"]
                        ws[f"E{f_deb}"].fill = fill_row_deb
                        ws[f"E{f_deb}"].alignment = Alignment(horizontal='center')
                        ws[f"E{f_deb}"].border = thin_border

                        ws[f"F{f_deb}"] = str(r["Descripcion"])
                        ws[f"F{f_deb}"].fill = fill_row_deb
                        ws[f"F{f_deb}"].border = thin_border

                        ws[f"G{f_deb}"] = r["Importe"]
                        ws[f"G{f_deb}"].number_format = '"$ "#,##0.00'
                        ws[f"G{f_deb}"].fill = fill_row_deb
                        ws[f"G{f_deb}"].border = thin_border
                        f_deb += 1
                    
                    # Total Débitos
                    ws.merge_cells(f"E{f_deb}:F{f_deb}")
                    ws[f"E{f_deb}"] = "TOTAL DÉBITOS"
                    ws[f"E{f_deb}"].font = Font(bold=True)
                    ws[f"E{f_deb}"].alignment = Alignment(horizontal='right')
                    ws[f"E{f_deb}"].fill = fill_col_deb
                    ws[f"E{f_deb}"].border = thin_border
                    
                    ws[f"G{f_deb}"] = f"=SUM(G{start_d}:G{f_deb-1})"
                    ws[f"G{f_deb}"].number_format = '"$ "#,##0.00'
                    ws[f"G{f_deb}"].font = Font(bold=True)
                    ws[f"G{f_deb}"].fill = fill_col_deb
                    ws[f"G{f_deb}"].border = thin_border
                    f_deb += 1

                f_ini = "B3"
                f_tot_cred = f"C{f_cred-1}" if not creditos.empty else "0"
                f_tot_deb = f"G{f_deb-1}" if not debitos.empty else "0"
                f_fin = "B4"
                
                # Asignamos a D7
                ws["D7"] = f"={f_ini}+{f_tot_cred}-{f_tot_deb}-{f_fin}"
                ws["D7"].number_format = '"$ "#,##0.00'

                # FORMATO CONDICIONAL: ROJO SI NO ES CERO
                red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                red_font = Font(color='9C0006', bold=True)
                
                ws.conditional_formatting.add('D7', 
                    CellIsRule(operator='notEqual', formula=['0'], stopIfTrue=True, fill=red_fill, font=red_font))

                # Ajustar ancho de columnas
                ws.column_dimensions["A"].width = 12
                ws.column_dimensions["B"].width = 45
                ws.column_dimensions["C"].width = 18
                ws.column_dimensions["D"].width = 25 
                ws.column_dimensions["E"].width = 12
                ws.column_dimensions["F"].width = 45
                ws.column_dimensions["G"].width = 18

                # Guardar en BytesIO
                wb.save(output)

                # Preparar el archivo para descarga
                output.seek(0)

                st.success(f"Archivo Excel creado con {len(movimientos)} movimientos")
                return output.getvalue()

            except Exception as e:
                st.error(f"Error creando archivo Excel: {str(e)}")
                return None
        else:
            falta = []
            if not saldo_inicial:
                falta.append("inicial")
            if not saldo_final:
                falta.append("final")
            menciones = re.findall(r"(?i)saldo[^\n]{0,60}", texto)[:5]
            st.warning(
                f"No se encontró saldo {' y '.join(falta)}. "
                f"Primeras menciones de 'Saldo' en el PDF extraído: {menciones}"
            )
            return None

    except Exception as e:
        st.error(f"Error al procesar el archivo: {str(e)}")
        import traceback

        st.error(f"Detalles del error: {traceback.format_exc()}")
        return None
